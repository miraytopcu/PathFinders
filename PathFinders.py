import json
import math
from geopy.distance import geodesic
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from time import sleep
import random
import openpyxl

# ---------------------------------------------------------------------------------------------------------------
class Ulasim:
    def __init__(self, from_city, to_city):
        self.sehirler = []  # Şehir listesini tutmak için boş bir liste oluşturduk
        self.from_city = from_city
        self.to_city = to_city

    # @staticmethod
    def distance(self):
        # JSON dosyasından verileri çekme
        with open('cities_of_turkey.json', 'r', encoding='utf-8') as file:
            veri = json.load(file)

        # Şehir isimlerini büyük harfle başlayıp kalanı küçük olacak şekilde düzenleme
        sehirler = [self.from_city.capitalize() , self.to_city.capitalize()]
        if sehirler[0] == "Istanbul":
            sehirler[0] = sehirler[0].replace("I", "İ")
        elif sehirler[1] == "Istanbul":
            sehirler[1] = sehirler[1].replace("I", "İ")
        elif sehirler[0] == "Izmir":
            sehirler[0] = sehirler[0].replace("I", "İ")
        elif sehirler[1] == "Izmir":
            sehirler[1] = sehirler[1].replace("I", "İ")
        lats = []   # Girilen inputtaki şehrin enlem bilgisi
        longs = []  # Girilen inputtaki şehrin boylam bilgisi
        for sehir in sehirler:
            found = False
            for data in veri:
                if data["name"] == sehir:
                    lat = data['latitude']
                    long = data['longitude']
                    lats.append(lat)
                    longs.append(long)
                    found = True
                    break
            if not found:
                print(f"{sehir} adlı şehir veri setinde bulunamadı.")
        
        if len(lats) >= 2 and len(longs) >= 2:        # Listede en az iki şehir olduğunu kontrol etmek için
            coord1 = (lats[0], longs[0])
            coord2 = (lats[1], longs[1])
            return int(geodesic(coord1, coord2).kilometers)    # Geodesic kütüphanesi iki koordinat arası mesafeyi hesaplar
        else:
            return "İki şehir arasında mesafe hesaplanamadı."

# -------------------------------------------------------------------------------------------------------------

    def otobus(self):
        distance = self.distance()
        koltuk_sayisi = 40

        def personel_ucreti():
            sofor = distance * 1.5 * 2     # 2 sofor var, km başı 1.5 tl
            gorevli = distance * 1
            return sofor + gorevli

        def ikram_servis():
            ikram_maliyeti = koltuk_sayisi * 30
            servis_maliyeti = koltuk_sayisi * 40
            return ikram_maliyeti + servis_maliyeti

        def otogar_masrafi():
            durulan_otogar_sayisi = math.floor(distance / 150)     # 150 kmde bir otogarda duruyor ve her durduğu otogarda 450 harcıyor
            return durulan_otogar_sayisi * 450

        def bakimlar():
            yillik_kasko = 400000
            sigorta = 80000
            tek_lastik = 15000
            lastikler = tek_lastik * 6
            yilda_gidilen_km = 330000
            km_basina_bakim_masraflari = (yillik_kasko + sigorta + lastikler) / yilda_gidilen_km
            sefer_basi_bakim_masrafi = km_basina_bakim_masraflari * distance
            return sefer_basi_bakim_masrafi

        def yakit_masrafi():       # Mazot fiyatları sürekli güncellendiği için seleniumdan xpath kullanarak veriyi aldık
            chromeOptions = Options()
            chromeOptions.add_argument("--incognito")
            chromeOptions.add_argument("--headless")
            driver = webdriver.Chrome(options=chromeOptions)
            driver.get("https://www.aytemiz.com.tr/akaryakit-fiyatlari/motorin-fiyatlari")
            driver.implicitly_wait(5)     # sitenin açılması için bekleme süresi

            try:
                fiyat_bilgisi = driver.find_element("xpath", '//*[@id="fuel-price-table"]/tbody/tr[6]/td[3]').text
                fiyat = float(fiyat_bilgisi.replace(',', '.'))
                litre = distance * 0.28
                return litre * fiyat
            except Exception as e:
                print(f"Hata oluştu: {e}")
            
            driver.quit()  

        def diger_masraflar():
            mtv = 13500     # motorlu taşıtlar vergisi 
            aylik_kasko = 120000
            yilda_gidilen_km = 330000
            km_basina_kasko = (aylik_kasko * 12 ) / yilda_gidilen_km
            sefer_basina_kasko_mtv = (distance * km_basina_kasko) + ((mtv / yilda_gidilen_km) * distance)
            return sefer_basina_kasko_mtv
        
        toplam_masraf = personel_ucreti() + ikram_servis() + otogar_masrafi() + bakimlar() + yakit_masrafi() + diger_masraflar()
        amortisman = toplam_masraf     # kaarda amortisman içinde
        return int((toplam_masraf + amortisman) / koltuk_sayisi)
    
# --------------------------------------------------------------------------------------------------------------
    
    def load_cities_and_regions(self):
        # Load cities and regions from the JSON file
        with open('cities_of_turkey.json', 'r', encoding='utf-8') as file:
            veri = json.load(file)
        
        self.sehirler = [data["name"] for data in veri]
        self.sehir_regionleri = {data["name"]: data["region"] for data in veri}
    
    def ucak(self):
        # JSON dosyasından verileri çekme
        with open('cities_of_turkey.json', 'r', encoding='utf-8') as file:
            veri = json.load(file)

        # Kullanıcının girdiği şehir isimlerini al
        sehirler = [self.from_city.capitalize() , self.to_city.capitalize()]
        if sehirler[0] == "Istanbul":
            sehirler[0] = sehirler[0].replace("I", "İ")
        elif sehirler[1] == "Istanbul":
            sehirler[1] = sehirler[1].replace("I", "İ")
        elif sehirler[0] == "Izmir":
            sehirler[0] = sehirler[0].replace("I", "İ")
        elif sehirler[1] == "Izmir":
            sehirler[1] = sehirler[1].replace("I", "İ")

        # Ensure cities and regions data is loaded
        if not self.sehirler or not self.sehir_regionleri:
            self.load_cities_and_regions()

        # Bölgeler
        regions = {
            "Ege": "İzmir",
            "Akdeniz": "Antalya",
            "Güneydoğu Anadolu": "Gaziantep",
            "İç Anadolu": "Ankara",
            "Doğu Anadolu": "Malatya",
            "Marmara": "İstanbul",
            "Karadeniz": "Samsun"
        }

        # ucak_bileti = 0

        for data in veri:
            if data["name"] == sehirler[0]:
                hava1 = data["havaalani"]
            if data["name"] == sehirler[1]:
                hava2 = data["havaalani"]

        if hava1 == "-" and hava2 == "-":
            bolge1 = self.sehir_regionleri.get(sehirler[0], "Bilinmeyen Bölge")
            sehir_bolge1 = regions.get(bolge1, "Bilinmeyen Bölge")
            bolge2 = self.sehir_regionleri.get(sehirler[1], "Bilinmeyen Bölge")
            sehir_bolge2 = regions.get(bolge2, "Bilinmeyen Bölge")
            print(f"{sehirler[0]} adlı şehirde havaalanı bulunmamaktadır! {sehir_bolge1} şehrinden uçağa binebilirsiniz.")
            print(f"{sehirler[1]} adlı şehirde havaalanı bulunmamaktadır! {sehir_bolge2} şehrinden uçağa binebilirsiniz.")
            secim = input(f"Aktarmalı gitmek isterseniz 'e', yolculuğu otobüsle yapmak isterseniz 'h' tuşlayınız: ")
            if secim == "h":
                otobus_fiyat = self.otobus()
                return otobus_fiyat
            elif secim == "e":
                self.from_city = sehirler[0]
                self.to_city = sehir_bolge1
                otobus_fiyat1 = self.otobus()
                self.from_city = sehir_bolge1
                self.to_city = sehir_bolge2
                distance = self.distance()
                ucak_fiyat = math.ceil(distance * 3)
                self.from_city = sehir_bolge2
                self.to_city = sehirler[1]
                otobus_fiyat2 = self.otobus()
                return otobus_fiyat1 + ucak_fiyat + otobus_fiyat2
            else:
                return f"Tercihini gözden geçir!"

        elif hava1 == "+" and hava2 == "+":
            distance = self.distance()  # Mesafeyi hesapla
            ucak_bileti = math.ceil(distance * 3)  # km başına 3 TL
            return ucak_bileti
        
        elif hava1 == "-" and hava2 == "+":
            bolge = self.sehir_regionleri.get(sehirler[0], "Bilinmeyen Bölge")
            sehir_bolge = regions.get(bolge, "Bilinmeyen Bölge")
            print(f"{sehirler[0]} adlı şehirde havaalanı bulunmamaktadır! {sehir_bolge} şehrinden uçağa binebilirsiniz.")
            secim = input(f"Aktarmalı gitmek isterseniz 'e', yolculuğu otobüsle yapmak isterseniz 'h' tuşlayınız: ")
            if secim == "h":
                otobus_fiyat = self.otobus()
                return otobus_fiyat
            elif secim == "e":
                self.from_city = sehirler[0]
                self.to_city = sehir_bolge
                otobus_fiyat = self.otobus()
                self.from_city = sehir_bolge
                self.to_city = sehirler[1]
                distance = self.distance()
                ucak_fiyat = math.ceil(distance * 3)
                return otobus_fiyat + ucak_fiyat
            else:
                return f"Tercihini gözden geçir!"
            
        elif hava1 == "+" and hava2 == "-":
            bolge = self.sehir_regionleri.get(sehirler[1], "Bilinmeyen Bölge")
            sehir_bolge = regions.get(bolge, "Bilinmeyen Bölge")
            print(f"{sehirler[1]} adlı şehirde havaalanı bulunmamaktadır! {sehir_bolge} şehrinden uçağa binebilirsiniz.")
            secim = input(f"Aktarmalı gitmek isterseniz 'e', yolculuğu otobüsle yapmak isterseniz 'h' tuşlayınız: ")
            if secim == "h":
                otobus_fiyat = self.otobus()
                return otobus_fiyat
            elif secim == "e":
                self.from_city = sehirler[0]
                self.to_city = sehir_bolge
                distance = self.distance()
                ucak_fiyat = math.ceil(distance * 3)
                self.from_city = sehir_bolge
                self.to_city = sehirler[1]
                otobus_fiyat = self.otobus()
                return otobus_fiyat + ucak_fiyat
            else:
                return f"Tercihini gözden geçir!"
               
        else:
            return f"Tercihini gözden geçir!!!"
        
# ------------------------------------------------------------------------------------------------------
class Otel:
    def __init__(self, location):
        self.location = location
    
    def hotels(self):
        if self.location is None:
            print("Location is not specified.")
            return
        # Excel dosyasının yolunu belirtin
        dosya_yolu = 'PathFinders_oteller.xlsx'

        # Excel dosyasını ve sayfasını yükleyin
        kitap = openpyxl.load_workbook(dosya_yolu)
        sayfa = kitap.active  # Aktif sayfa seçili hale gelir

        self.location = self.location.capitalize()
        if self.location == "Istanbul":
            self.location = self.location.replace("I", "İ")
        elif self.location == "Izmir":
            self.location = self.location.replace("I", "İ")

        # Şehir için uygun otelleri ve fiyatlarını depolamak için liste oluşturun
        oteller = []

        # Satır sayısını belirleyin
        satir_sayisi = sayfa.max_row

        # Verileri çekmek için bir döngü yazın
        for i in range(2, satir_sayisi + 1):  # Satırlar arasında döngü (başlık satırı atlanarak)
            sehir = sayfa.cell(i, 2).value  # 2. sütundaki şehir ismini al
            
            # Şehir eşleşiyorsa otelleri al
            if sehir.lower() == self.location.lower():
                for j in range(3, 6):  # 3. sütundan 5. sütuna kadar oteller
                    otel = sayfa.cell(i, j).value  # Otel ismini al
                    fiyat = sayfa.cell(i, j + 3).value  # Aynı otelin fiyatını al (3 sütun sonrası)
                    
                    # Otel ve fiyatı listeye ekle
                    oteller.append((otel, fiyat))
                
                # Kamp alanı ve fiyatını al
                kamp_alani = sayfa.cell(i, 9).value  # 9. sütundaki kamp alanı ismini al
                kamp_fiyati = sayfa.cell(i, 10).value  # 10. sütundaki kamp fiyatını al
                
                # Kamp alanını da listeye ekle
                oteller.append((kamp_alani, kamp_fiyati))  # Kamp alanları için adres ve telefon yoksa boş bırakın
            
            # İlgili şehir için 3 otel ve 1 kamp alanı bulduysak döngüden çık
            if len(oteller) >= 4:
                break

        # Şehir için otel ve kamp seçenekleri bulunduğunda sonucu yazdır
        if oteller:
            print(f"{self.location} için mevcut otel ve kamp seçenekleri:")
            for idx, (otel, fiyat) in enumerate(oteller[:4], start=1):  # İlk 3 otel ve 1 kamp alanını yazdır
                print(f"{idx}. seçenek: Otel/Kamp: {otel} \nFiyat: {fiyat} TL")
            
            # Kullanıcıdan seçim yapmasını iste
            while True:
                try:
                    secim = int(input("Hangi oteli/kampı seçmek istersiniz (1-4): "))
                    if 1 <= secim <= len(oteller[:4]):
                        secili_otel, secili_fiyat = oteller[secim - 1]
                        # secili_fiyat değerini sayıya dönüştürün
                        secili_fiyat = float(secili_fiyat)  # Eğer fiyatlar tamsayı ise int() kullanın
                        print(f"Seçtiğiniz Otel/kamp: {secili_otel} \nFiyat: {secili_fiyat} TL")
                        
                        return secili_fiyat
                    else:
                        print("Lütfen 1 ile 4 arasında bir sayı girin.")
                except ValueError:
                    print("Lütfen geçerli bir sayı girin.")
        else:
            print(f"{self.location} için mevcut otel/kamp bulunamadı.")

# ---------------------------------------------------------------------------------------------------------

def kampanya(location):
    hizmet_bedeli = 200

    with open('cities_of_turkey.json', 'r', encoding='utf-8') as file:
        veri = json.load(file)

    sehirler_listesi = []
    for data in veri:
        sehirler_listesi.append(data["name"])

    rastgele_sehirler = random.sample(sehirler_listesi, 15)
    # print(f"Rastgele belirlenen şehirler: {', '.join(rastgele_sehirler)}")

    # Kullanıcı şehri rastgele belirlenen şehirler arasında mı kontrol et
    if location in rastgele_sehirler:
        print(f"{location} seçildiği için %15 indirim uygulanacak!")
        indirimli_fiyat = hizmet_bedeli * 0.85
        # print(f"İndirimli toplam fiyat: {indirimli_fiyat:.2f} TL")
        return indirimli_fiyat
    else:
        print(f"{location} kampanyalı şehirler arasında değil, indirim uygulanmayacak.")
        return hizmet_bedeli

# ------------------------------------------------------------------------------------------------------

while True:        
    ad_soyad = input("Adınız Soyadınız: ")
    # Kullanıcıyı karşılama mesajı
    print(f"\nMerhaba {ad_soyad}, PathFinders'a hoş geldiniz!")
    print(f"Size uygun tatil önerileri sunacağız.\n")

    num_person = int(input("Katılacak kişi sayısını giriniz: "))
    num_day = int(input("Kalınacak gece sayısını giriniz: "))
    max_butce = float(input("Tatil için ayırabileceğiniz max bütçeyi giriniz: "))
    rehber = input("Rehber hizmeti istiyorsanız 'e' istemiyorsanız 'h' tuşlayınız: ")

    if rehber == "e":
        rehber_fiyat = 500
    else:
        rehber_fiyat = 0
    # İşlem türü seçimi
    print("Lütfen yapmak istediğiniz işlemi seçin:")
    print("1) Tatil yapmak istediğim şehri ben seçmek istiyorum.")
    print("2) Bir kategoriye göre tatil yapmak istiyorum (Kültürel, Lezzet, Deniz, Doğa)")
    islem = input("\nSeçiminizi yapın (1, 2): ")

    # -------------------------------------------------------------------------------------------------------------

    if islem == '1':
        print("\nSeçtiniz: Tatil yapmak istediğim şehri ben seçmek istiyorum.")
        # Bu seçeneğe göre yapılacak işlemler burada tanımlanacak.
        tofrom = input("Sırasıyla çıkmak ve gitmek istediğiniz illeri arada boşluk karakteri kullanarak giriniz: ")
        tofromx = tofrom.split()
        from_city = tofromx[0]
        to_city = tofromx[1]
        nesne_ulasim = Ulasim(from_city, to_city)
        nesne_otel = Otel(to_city)

        ulasim_tercih = input("Otobüsle yolculuk yapmak istiyorsanız 'o' uçakla yolculuk yapmak istiyorsanız 'u' tuşlayınız: ")
        if ulasim_tercih == "o":
            # invoke otobus
            bilet_fiyati = nesne_ulasim.otobus()
        elif ulasim_tercih == "u":
            # invoke ucak
            bilet_fiyati = nesne_ulasim.ucak()
        print(f"Bilet Fiyatı: {bilet_fiyati}\n")
        # otel vs
        otel_fiyati = nesne_otel.hotels()
        hizmet = kampanya(to_city)   

    # --------------------------------------------------------------------------------------------------

    elif islem == '2':
        print("\nSeçtiniz: Bir kategoriye göre tatil yapmak istiyorum")
        # Bu seçeneğe göre yapılacak işlemler burada tanımlanacak.
        print("Lütfen seyahat türünüzü seçin:")
        print("1. Deniz Tatili")
        print("2. Lezzet Turu")
        print("3. Kültürel Gezi")
        print("4. Doğa Tatili")

        secim = input("Seçiminizi girin (1-4): ")

        with open('cities_of_turkey.json', 'r', encoding='utf-8') as file:
            veri = json.load(file)
        list_deniz = []
        list_lezzet = []
        list_kultur = []
        list_doga = []

        for i in veri:
            if i["kategori"] == "Deniz":
                list_deniz.append(i["name"])
            elif i["kategori"] == "Lezzet":
                list_lezzet.append(i["name"])
            elif i["kategori"] == "Kültürel":
                list_kultur.append(i["name"])
            elif i["kategori"] == "Doğa":
                list_doga.append(i["name"])
            else:
                print("Geçersiz seçim! Lütfen tekrar deneyin.")

        if secim == '1':
            kategori = "Deniz"
            randoms = random.sample(list_deniz, min(5, len(list_deniz)))  # Ensure max sample size is limited by list size
        elif secim == '2':
            kategori = "Lezzet"
            randoms = random.sample(list_lezzet, min(5, len(list_lezzet)))
        elif secim == '3':
            kategori = "Kültürel"
            randoms = random.sample(list_kultur, min(5, len(list_kultur)))
        elif secim == '4':
            kategori = "Doğa"
            randoms = random.sample(list_doga, min(5, len(list_doga)))
        else:
            print("Geçersiz seçim! Lütfen tekrar deneyin.")
            break

        print(f"{kategori} kategorisi için size önerdiğimiz seçenekleri aşağıda ulaşabilirsiniz.\n")
        for x in randoms:
            print(x)
        sehir_secim = input("Lütfen sizlere sunduğumuz seçenekler arasından tercih yapınız: ")
        sehir_secim = sehir_secim.capitalize()
        if sehir_secim == "Istanbul":
            sehir_secim = sehir_secim.replace("I", "İ")
        elif sehir_secim == "Istanbul":
            sehir_secim = sehir_secim.replace("I", "İ")
        elif sehir_secim == "Izmir":
            sehir_secim = sehir_secim.replace("I", "İ")
        elif sehir_secim == "Izmir":
            sehir_secim = sehir_secim.replace("I", "İ")

        cikis_sehir = input("Lütfen çıkış yapacağınız ili giriniz: ")
        cikis_sehir = cikis_sehir.capitalize()
        if cikis_sehir == "Istanbul":
            cikis_sehir = cikis_sehir.replace("I", "İ")
        elif cikis_sehir == "Istanbul":
            cikis_sehir = cikis_sehir.replace("I", "İ")
        elif cikis_sehir == "Izmir":
            cikis_sehir = cikis_sehir.replace("I", "İ")
        elif cikis_sehir == "Izmir":
            cikis_sehir = cikis_sehir.replace("I", "İ")

        nesne_ulasim = Ulasim(cikis_sehir, sehir_secim)
        ulasim_tercih = input("Otobüsle yolculuk yapmak istiyorsanız 'o' uçakla yolculuk yapmak istiyorsanız 'u' tuşlayınız: ")
        if ulasim_tercih == "o":
            # invoke otobus
            bilet_fiyati = nesne_ulasim.otobus()
        elif ulasim_tercih == "u":
            # invoke ucak
            bilet_fiyati = nesne_ulasim.ucak()
        print(f"Bilet Fiyatı: {bilet_fiyati}\n")

        nesne_otel = Otel(sehir_secim)
        otel_fiyati = nesne_otel.hotels()
        hizmet = kampanya(sehir_secim)

    else:
        print("\nGeçersiz seçim. Lütfen tekrar deneyin.")
        break

    # ------------------------------------------------------------------------------------------------------------
    # Toplam tutar tatil planları kodu yazıldıktan sonra oradan çekilebilir.
    toplam_tutar = int((bilet_fiyati * num_person) + hizmet + rehber_fiyat + (otel_fiyati * num_day * num_person))

    if toplam_tutar > max_butce:
        print(f"Seçimlerinizin toplam fiyatı max bütçenizden {toplam_tutar - max_butce} TL kadar fazladır!")
        tercih = input("İşlemi onaylayıp ödeme kısmına geçmek istiyorsanız lütfen 'e' tuşlayınız, diğer seçimlerinizde işleminiz sonlandırılacaktır! ")

        if tercih != "e":
            print("İşleminiz sonlandırılıyor.")
            break

    # Kullanıcıya ödeme seçeneklerini sunma
    print("Lütfen ödeme yöntemini seçin:")
    print("1) Kredi Kartı (%2 komisyonlu)")
    print("2) Banka Kartı")
    print("3) Nakit Ödeme")
    print("4) Havale ile Ödeme (%5 indirimli)")

    # Ödeme türü seçimi
    odeme = input("\nSeçiminizi yapın (1, 2, 3 veya 4): ")

    if odeme == '1':
        komisyon_orani = 0.02
        komisyonlu_tutar = toplam_tutar * (1 + komisyon_orani)
        print(f"\nSeçtiniz: Kredi Kartı ile Ödeme")
        print(f"Ödenecek Tutar (komisyonlu): {komisyonlu_tutar:.2f} TL")
        break
            
    elif odeme == '2':
        print(f"\nSeçtiniz: Banka Kartı ile Ödeme")
        print(f"Ödenecek Tutar: {toplam_tutar:.2f} TL")
        break
            
    elif odeme == '3':
        print(f"\nSeçtiniz: Nakit Ödeme")
        print(f"Ödenecek Tutar: {toplam_tutar:.2f} TL")
        break
            
    elif odeme == '4':
        indirimli_tutar = toplam_tutar * 0.95
        print(f"\nSeçtiniz: Havale ile Ödeme")
        print(f"Ödenecek Tutar (indirimli): {indirimli_tutar:.2f} TL")
        break
            
    else:
        print("\nGeçersiz seçim. Lütfen tekrar deneyin.")
        break

