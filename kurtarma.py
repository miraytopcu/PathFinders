# %%
import openpyxl

# %%
dosya= openpyxl.load_workbook("./PathFinders_oteller.xlsx")
print(dosya.sheetnames)

# %%
print("aktif sayfa:" + dosya.active.title)

# %%
sayfa= dosya["otel"]
deger= sayfa["B4"].value
print(deger)
#alternatif
deger=dosya["otel"]["B4"].value
print(deger)

# %%
sayfa=dosya["otel"]
veri= sayfa.cell(3,2).value
print("veri:" +str(veri)) #str sayıyı dönüştürmek için

# %%
satir_sayisi=sayfa.max_row
sutun_sayisi=sayfa.max_column

# %%
for i in range(2,satir_sayisi):
    for j in range (2,sutun_sayisi):
        veri=sayfa.cell(i,j).value
        print(veri)

# %%
def excel_listeler_listesine_cevir(dosyayeri,sayfaadi):
    dosya=openpyxl.load_workbook(dosyayeri)
    sayfa=dosya[sayfaadi]
    satir_sayisi=sayfa.max_row
    sutun_sayisi=sayfa.max_column
    data=[]
    for i in range(2,satir_sayisi):
        satir=[]
        for j in range(1,sutun_sayisi):
            satir.append(sayfa.cell(i,j).value)
        data.append(satir)
    return data

tumveri=excel_listeler_listesine_cevir("./PathFinders_oteller.xlsx","otel")

for i in range(len(tumveri)):
    liste=tumveri[i]
    sıra_no=liste[0]
    iller=liste[1]
    otel1=liste[2]
    otel2=liste[3]
    otel3=liste[4]
    fiyat1=liste[5]
    fiyat2=liste[6]
    fiyat3=liste[7]
        

# %%
def excel_listeler_listesine_cevir(dosyayeri,sayfaadi):
    dosya=openpyxl.load_workbook(dosyayeri)
    sayfa=dosya[sayfaadi]
    satir_sayisi=sayfa.max_row
    sutun_sayisi=sayfa.max_column
    data=[]
    for i in range(2,satir_sayisi):
        satir=[]
        for j in range(2,sutun_sayisi):
            satir.append(sayfa.cell(i,j).value)
        data.append(satir)
    return data
tumveri=excel_listeler_listesine_cevir("./PathFinders_oteller.xlsx","otel")
print(tumveri)

# %%
import pandas as pd
dosya= openpyxl.load_workbook("./PathFinders_oteller.xlsx")
df = pd.read_excel("./PathFinders_oteller.xlsx","otel")  

cells_to_fetch = [
    ('B2', 'C2', 'F2'),
    ('B2', 'D2', 'G2'),
    ('B2', 'E2', 'H2')
]

# Hücrelerdeki verileri çekmek için bir fonksiyon yazın
def excel_listeler_listesine_cevir(dosyayeri,sayfaadi):
    dosya=openpyxl.load_workbook(dosyayeri)
    sayfa=dosya[sayfaadi]
    satir_sayisi=sayfa.max_row
    sutun_sayisi=sayfa.max_column
    data=[]
    for i in range(2,satir_sayisi):
        satir=[]
        for j in range(2,sutun_sayisi):
            satir.append(sayfa.cell(i,j).value)
        data.append(satir)
    return data

# Verileri çekin
tumveri=excel_listeler_listesine_cevir("./PathFinders_oteller.xlsx","otel")

# Verileri yazdırın
for data in tumveri:
    print(data)


# %%
for i in range(2, satir_sayisi, 2):  # İki satır atlayarak ilerlemek için step=2
    for j in range(2, sutun_sayisi):  # Sütunlar için adım normal kalıyor
        veri = sayfa.cell(i, j).value
        print(veri)


# %%
import openpyxl

# Excel dosyasının yolunu belirtin
dosya_yolu = 'PathFinders_oteller.xlsx'

# Excel dosyasını ve sayfasını yükleyin
kitap = openpyxl.load_workbook(dosya_yolu)
sayfa = kitap.active  # Aktif sayfa seçili hale gelir

# Satır sayısını belirtin
satir_sayisi = sayfa.max_row

# Verileri çekmek için bir döngü yazın
for i in range(2, satir_sayisi + 1):  # Satırlar arasında döngü (başlık satırı atlanarak)
    sehir = sayfa.cell(i, 2).value  # 2. sütundaki şehir ismini al
    
    # Her otel ve fiyatı eşleştir
    for j in range(3, 6):  # 3. sütundan 5. sütuna kadar oteller
        otel = sayfa.cell(i, j).value  # Otel ismini al
        fiyat = sayfa.cell(i, j + 3).value  # Aynı otelin fiyatını al (3 sütun sonrası)
        
        # Sonucu yazdır
        print(f"{sehir} - {otel} - {fiyat}")





